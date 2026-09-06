from __future__ import annotations

import io
from decimal import Decimal, InvalidOperation
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from pydantic import BaseModel

from app.services.excel_spec import SPECS, Column

HEADER_FILL = PatternFill("solid", fgColor="E8EDF3")
KEY_FILL = PatternFill("solid", fgColor="FFF3D6")


class ImportRowError(BaseModel):
    row: int
    message: str


class ImportResult(BaseModel):
    created: int = 0
    updated: int = 0
    skipped: int = 0
    errors: list[ImportRowError] = []


# ---------------------------------------------------------------------------
# Writing
# ---------------------------------------------------------------------------
def build_workbook(kind: str, rows: list[dict[str, Any]] | None = None) -> bytes:
    """A workbook for `kind`, empty (a blank template) or filled with rows.

    Row 1 is the headers. Row 2 is the guidance line - notes explaining what a
    column means or when it is required - frozen along with the header so it
    stays visible while scrolling. Data starts at row 3.
    """
    columns, _, title = SPECS[kind]
    wb = Workbook()
    ws = wb.active
    ws.title = title

    for i, col in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=i, value=col.header)
        cell.font = Font(bold=True)
        cell.fill = KEY_FILL if col.key else HEADER_FILL
        cell.alignment = Alignment(vertical="center", wrap_text=True)

        note = ws.cell(row=2, column=i, value=col.note or "")
        note.font = Font(italic=True, size=9, color="666666")
        note.alignment = Alignment(vertical="top", wrap_text=True)

        width = max(len(col.header) + 4, 14)
        ws.column_dimensions[get_column_letter(i)].width = min(width, 32)

    ws.freeze_panes = "A3"
    ws.row_dimensions[2].height = 28

    for r, row in enumerate(rows or [], start=3):
        for i, col in enumerate(columns, start=1):
            ws.cell(row=r, column=i, value=_to_cell(row.get(col.field)))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _to_cell(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, (int, float, str)):
        return value
    return str(value)


# ---------------------------------------------------------------------------
# Reading
# ---------------------------------------------------------------------------
def parse_workbook(kind: str, content: bytes) -> tuple[list[dict[str, Any]], list[ImportRowError]]:
    """Read an uploaded workbook into dicts keyed by schema field name.

    Matching is by header text, case- and whitespace-insensitive, so a file
    with reordered or extra columns still works and a missing optional column
    simply means that field is not set. Unknown columns are ignored rather
    than rejected - people add working notes to these sheets.
    """
    columns, _, _ = SPECS[kind]
    try:
        wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    except Exception as exc:  # noqa: BLE001 - surfaced to the user as a row error
        return [], [ImportRowError(row=0, message=f"Could not read the file: {exc}")]

    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)

    try:
        header_row = next(rows_iter)
    except StopIteration:
        return [], [ImportRowError(row=0, message="The sheet is empty")]

    by_header = {_norm(c.header): c for c in columns}
    index_to_col: dict[int, Column] = {}
    for idx, raw in enumerate(header_row):
        col = by_header.get(_norm(raw))
        if col is not None:
            index_to_col[idx] = col

    if not index_to_col:
        return [], [
            ImportRowError(
                row=1,
                message="No recognised column headers. Download the template and use its header row.",
            )
        ]

    parsed: list[dict[str, Any]] = []
    errors: list[ImportRowError] = []

    for row_number, raw_row in enumerate(rows_iter, start=2):
        if _is_blank(raw_row):
            continue
        # row 2 of the template is the guidance line, not data
        if row_number == 2 and _looks_like_guidance(raw_row, columns):
            continue

        record: dict[str, Any] = {"__row__": row_number}
        row_errors: list[str] = []
        for idx, col in index_to_col.items():
            raw = raw_row[idx] if idx < len(raw_row) else None
            try:
                record[col.field] = _coerce(raw, col)
            except ValueError as exc:
                row_errors.append(f"{col.header}: {exc}")

        if row_errors:
            errors.append(ImportRowError(row=row_number, message="; ".join(row_errors)))
        else:
            parsed.append(record)

    return parsed, errors


def _norm(value: Any) -> str:
    return " ".join(str(value or "").split()).casefold()


def _is_blank(row: tuple[Any, ...]) -> bool:
    return all(c is None or (isinstance(c, str) and not c.strip()) for c in row)


def _looks_like_guidance(row: tuple[Any, ...], columns: list[Column]) -> bool:
    notes = {_norm(c.note) for c in columns if c.note}
    return any(_norm(cell) in notes for cell in row if cell)


def _coerce(raw: Any, col: Column) -> Any:
    if raw is None or (isinstance(raw, str) and not raw.strip()):
        return None
    if col.type == "text":
        return str(raw).strip()
    if col.type == "int":
        try:
            return int(float(str(raw).strip()))
        except (TypeError, ValueError):
            raise ValueError(f"{raw!r} is not a whole number") from None
    if col.type in ("decimal", "latlng"):
        try:
            return Decimal(str(raw).strip())
        except (TypeError, InvalidOperation):
            raise ValueError(f"{raw!r} is not a number") from None
    return str(raw).strip()
